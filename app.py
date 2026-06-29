import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CẤU HÌNH HỆ THỐNG & GIAO DIỆN
# =============================================================================
st.set_page_config(
    page_title="Hệ thống Quản lý Tiêu chí Chất lượng Vận hành Tín dụng",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Hệ thống Xử lý & Phân tích Dữ liệu Tín dụng")
st.markdown("---")

# Tạo Sidebar điều hướng giữa 4 tiêu chí
menu = st.sidebar.radio(
    "CHỌN TIÊU CHÍ XỬ LÝ DỮ LIỆU",
    [
        "Tiêu chí 1: Dư nợ cơ cấu",
        "Tiêu chí 2: Miễn giảm lãi",
        "Tiêu chí 3: Lũy kế nợ quá hạn",
        "Tiêu chí 4: Tín dụng đảm bảo bằng TSBD"
    ]
)

# -----------------------------------------------------------------------------
# CÁC HÀM TIỆN ÍCH CHUẨN HÓA LOGIC (MỚI)
# -----------------------------------------------------------------------------
def normalize_col(col):
    col = str(col).strip().upper()
    col = re.sub(r"\s+", "_", col)
    col = col.replace(".", "_")
    col = col.replace("-", "_")
    return col

def clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).strip().upper()

def clean_number(s):
    if pd.isna(s):
        return 0
    s = str(s).strip()
    if s == "":
        return 0
    s = s.replace(",", "")
    s = s.replace(" ", "")
    s = s.replace("(", "-").replace(")", "")
    return pd.to_numeric(s, errors="coerce")

def safe_divide(a, b):
    return np.where(b != 0, a / b, 0)

# -----------------------------------------------------------------------------
# HÀM ĐỌC FILE THÔNG MINH (CHẤP NHẬN XLSX & XLS, TRÁNH LỖI STRUCT.ERROR / HTML)
# -----------------------------------------------------------------------------
def read_excel_smart(uploaded_files, target_month=None, auto_normalize_cols=False):
    if not uploaded_files:
        return pd.DataFrame()
    
    if not isinstance(uploaded_files, list):
        uploaded_files = [uploaded_files]
        
    list_df = []
    for f in uploaded_files:
        f_bytes = f.read()
        f.seek(0)  # Reset con trỏ file về đầu
        df_temp = pd.DataFrame()
        
        try:
            # Thử cách 1: Dùng engine mặc định dựa vào đuôi file
            engine_pick = "xlrd" if f.name.endswith('.xls') else "openpyxl"
            df_temp = pd.read_excel(io.BytesIO(f_bytes), engine=engine_pick, dtype=str)
        except Exception:
            try:
                # Bước 2: Thử ép chéo cấu trúc engine đề phòng file sai đuôi mở rộng
                engine_pick = "openpyxl" if f.name.endswith('.xls') else "xlrd"
                df_temp = pd.read_excel(io.BytesIO(f_bytes), engine=engine_pick, dtype=str)
            except Exception:
                try:
                    # Bước 3: Thử đọc trực tiếp định dạng HTML (đối với file Core banking giả lập đuôi .xls)
                    df_list = pd.read_html(io.BytesIO(f_bytes))
                    df_temp = df_list[0].astype(str) if df_list else pd.DataFrame()
                except Exception as e:
                    st.error(f"❌ Không thể đọc được file: {f.name}. Lỗi chi tiết: {str(e)}")
                    continue
                    
        if not df_temp.empty:
            df_temp["SOURCE_FILE"] = f.name
            if target_month:
                df_temp["THANG"] = target_month
            
            if auto_normalize_cols:
                df_temp.columns = [normalize_col(c) for c in df_temp.columns]
            else:
                df_temp.columns = df_temp.columns.str.strip().str.upper()
                
            list_df.append(df_temp)
            
    if list_df:
        return pd.concat(list_df, ignore_index=True)
    return pd.DataFrame()

# Hàm bổ trợ để chuyển đổi DataFrame sang file Excel phục vụ nút Tải về thông thường
def to_excel_download(df, sheet_name="Sheet1"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


# =============================================================================
# TIÊU CHÍ 1: DƯ NỢ CƠ CẤU
# =============================================================================
if menu == "Tiêu chí 1: Dư nợ cơ cấu":
    st.header("📌 Tiêu chí 1: Đối chiếu & Phân bổ Dư nợ cơ cấu")
    st.info("💡 Hướng dẫn: Đăng tải đồng thời nhiều file mẫu CRM32 và CRM4 cho cả tháng này và tháng trước (Chấp nhận .xls và .xlsx).")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("🗓️ Dữ liệu Tháng này (T)")
        files_crm32_nay = st.file_uploader("Upload các file CRM32 Tháng này:", accept_multiple_files=True, key="t1_m32_n")
        files_crm4_nay = st.file_uploader("Upload các file CRM4 Tháng này:", accept_multiple_files=True, key="t1_m4_n")
    
    with col2:
        st.subheader("🗓️ Dữ liệu Tháng trước (T-1)")
        files_crm32_truoc = st.file_uploader("Upload các file CRM32 Tháng trước:", accept_multiple_files=True, key="t1_m32_t")
        files_crm4_truoc = st.file_uploader("Upload các file CRM4 Tháng trước:", accept_multiple_files=True, key="t1_m4_t")

    st.subheader("📂 Dữ liệu Phụ trợ")
    file_sddp = st.file_uploader("Upload file Số dư dự phòng DPRR_T5 (.xls, .xlsx)", key="t1_sddp")

    if st.button("🚀 Chạy Xử lý Tiêu chí 1"):
        if files_crm32_nay and files_crm4_nay and files_crm32_truoc and files_crm4_truoc and file_sddp:
            with st.spinner("Đang tính toán dư nợ cơ cấu..."):
                df_crm32_nay = read_excel_smart(files_crm32_nay)
                df_crm4_nay = read_excel_smart(files_crm4_nay)
                df_crm32_truoc = read_excel_smart(files_crm32_truoc)
                df_crm4_truoc = read_excel_smart(files_crm4_truoc)

                # --- LỚP BẢO VỆ CỘT TIÊU CHÍ 1 ---
                req_crm32 = ["SCHM_DESC", "CUSTSEQLN"]
                req_crm4 = ["CIF_KH_VAY", "DU_NO_PHAN_BO_QUY_DOI", "NHOM_NO", "KHACH_HANG"]
                
                err_32_nay = [c for c in req_crm32 if c not in df_crm32_nay.columns]
                err_4_nay = [c for c in req_crm4 if c not in df_crm4_nay.columns]
                err_32_trc = [c for c in req_crm32 if c not in df_crm32_truoc.columns]
                err_4_trc = [c for c in req_crm4 if c not in df_crm4_truoc.columns]
                
                if err_32_nay or err_4_nay or err_32_trc or err_4_trc:
                    st.error("❌ Phát hiện file upload thiếu cột hoặc dòng tiêu đề bị đặt sai vị trí!")
                    if err_32_nay: st.write(f"• File CRM32 Tháng này thiếu cột: `{err_32_nay}`")
                    if err_4_nay: st.write(f"• File CRM4 Tháng này thiếu cột: `{err_4_nay}`")
                    if err_32_trc: st.write(f"• File CRM32 Tháng trước thiếu cột: `{err_32_trc}`")
                    if err_4_trc: st.write(f"• File CRM4 Tháng trước thiếu cột: `{err_4_trc}`")
                    st.stop()

                # Xử lý tháng này
                df_crm32_cc_nay = df_crm32_nay[df_crm32_nay["SCHM_DESC"].astype(str).str.upper().str.replace(r"\s+", " ", regex=True).str.contains("CO CAU", na=False)].drop_duplicates(subset=["CUSTSEQLN"])
                df_join_nay = df_crm32_cc_nay.merge(df_crm4_nay, left_on="CUSTSEQLN", right_on="CIF_KH_VAY", how="inner")
                df_join_nay["DU_NO_PHAN_BO_QUY_DOI"] = pd.to_numeric(df_join_nay["DU_NO_PHAN_BO_QUY_DOI"], errors="coerce").fillna(0)
                
                df_kq_nay = df_join_nay.groupby(["CIF_KH_VAY", "SCHM_DESC","NHOM_NO"], as_index=False)["DU_NO_PHAN_BO_QUY_DOI"].sum().rename(columns={"DU_NO_PHAN_BO_QUY_DOI": "TONG_DU_NO_THANG_5","NHOM_NO":"NHOM_NO_T5"})
                df_kh = df_join_nay[["CIF_KH_VAY", "KHACH_HANG"]].drop_duplicates("CIF_KH_VAY")
                df_kq_nay = df_kq_nay.merge(df_kh, on="CIF_KH_VAY", how="left")

                # Xử lý tháng trước
                df_crm32_cc_truoc = df_crm32_truoc[df_crm32_truoc["SCHM_DESC"].astype(str).str.upper().str.replace(r"\s+", " ", regex=True).str.contains("CO CAU", na=False)].drop_duplicates(subset=["CUSTSEQLN"])
                df_join_truoc = df_crm32_cc_truoc.merge(df_crm4_truoc, left_on="CUSTSEQLN", right_on="CIF_KH_VAY", how="inner")
                df_join_truoc["DU_NO_PHAN_BO_QUY_DOI"] = pd.to_numeric(df_join_truoc["DU_NO_PHAN_BO_QUY_DOI"], errors="coerce").fillna(0)
                
                df_kq_truoc = df_join_truoc.groupby(["CIF_KH_VAY", "SCHM_DESC", "NHOM_NO"], as_index=False)["DU_NO_PHAN_BO_QUY_DOI"].sum().rename(columns={"DU_NO_PHAN_BO_QUY_DOI": "TONG_DU_NO_THANG_4","NHOM_NO":"NHOM_NO_T4"})
                df_kh_trc = df_join_truoc[["CIF_KH_VAY", "KHACH_HANG"]].drop_duplicates("CIF_KH_VAY")
                df_kq_truoc = df_kq_truoc.merge(df_kh_trc, on="CIF_KH_VAY", how="left")

                # Đối chiếu kết quả
                df_kq_nay_rename = df_kq_nay.rename(columns={"CIF_KH_VAY": "CIF_KH_VAY_NAY","SCHM_DESC": "SCHM_DESC_NAY", "KHACH_HANG":"TEN_KH_T5"})
                df_kq_truoc_rename = df_kq_truoc.rename(columns={"CIF_KH_VAY": "CIF_KH_VAY_TRUOC","SCHM_DESC": "SCHM_DESC_TRUOC", "KHACH_HANG":"TEN_KH_T4"})
                
                df_doi_chieu = df_kq_nay_rename.merge(df_kq_truoc_rename, left_on=["CIF_KH_VAY_NAY", "SCHM_DESC_NAY"], right_on=["CIF_KH_VAY_TRUOC", "SCHM_DESC_TRUOC"], how="outer").fillna(0)
                df_doi_chieu["CHENH_LECH"] = df_doi_chieu["TONG_DU_NO_THANG_5"] - df_doi_chieu["TONG_DU_NO_THANG_4"]
                df_doi_chieu = df_doi_chieu.sort_values("CHENH_LECH", ascending=False)

                # Đọc Map số dư dự phòng
                df_sddp = read_excel_smart(file_sddp)
                
                if "CIF" not in df_sddp.columns or "PHAT_SINH_NO" not in df_sddp.columns:
                    st.error("❌ File Phụ trợ SDDP phải chứa cột `CIF` và `PHAT_SINH_NO`!")
                    st.stop()
                    
                df_sddp["PHAT_SINH_NO"] = pd.to_numeric(df_sddp["PHAT_SINH_NO"], errors="coerce").fillna(0)
                
                df_doi_chieu["CIF_KH_VAY_NAY"] = df_doi_chieu["CIF_KH_VAY_NAY"].astype(str).str.strip().str.replace(".0","", regex=False)
                df_sddp["CIF"] = df_sddp["CIF"].astype(str).str.strip().str.replace(".0","", regex=False)
                
                df_sddp_sum = df_sddp.groupby("CIF")["PHAT_SINH_NO"].sum()
                df_doi_chieu["DPRR"] = df_doi_chieu["CIF_KH_VAY_NAY"].map(df_sddp_sum)

                st.success("🎉 Xử lý thành công!")
                st.dataframe(df_doi_chieu)

                excel_data = to_excel_download(df_doi_chieu, sheet_name="DU_NO_CO_CAU")
                st.download_button("📥 Tải File Kết Quả Excel", data=excel_data, file_name="DU_NO_CO_CAU_OUTPUT.xlsx")
        else:
            st.error("⚠️ Vui lòng cung cấp đầy đủ các file đầu vào bắt buộc!")

# ==========================================
# TIÊU CHÍ 2: MIỄN GIẢM LÃI
# ==========================================
elif menu == "Tiêu chí 2: Miễn giảm lãi":
    st.header("📌 Tiêu chí 2: Tổng hợp dữ liệu Miễn giảm lãi")
    
    col1, col2 = st.columns(2)
    with col1:
        file_hlawint = st.file_uploader("Upload file Màn hình HLAWINT (.xls, .xlsx)", key="t2_hl")
        file_noibang = st.file_uploader("Upload file Thu nợ Nội bảng (.xls, .xlsx)", key="t2_nb")
    with col2:
        file_ngoaibang = st.file_uploader("Upload file Thu nợ Ngoại bảng (.xls, .xlsx)", key="t2_ngb")
        file_thulai = st.file_uploader("Upload file Thu lãi (.xls, .xlsx)", key="t2_tl")

    if st.button("🚀 Chạy Xử lý Tiêu chí 2"):
        if file_hlawint and file_noibang and file_ngoaibang and file_thulai:
            with st.spinner("Đang tổng hợp dữ liệu..."):
                # 1. Khởi tạo dữ liệu từ HLAWINT
                df = read_excel_smart(file_hlawint)
                
                if not {"SOL_ID", "SOL_DESC", "CIF_ID", "CUST_NAME", "INTAMT_VND", "LY_DO"}.issubset(df.columns):
                    st.error("❌ File HLAWINT không đúng mẫu hoặc thiếu cột (SOL_ID, SOL_DESC, CIF_ID, CUST_NAME, INTAMT_VND, LY_DO)!")
                    st.stop()
                    
                cols = ['SOL_ID','SOL_DESC','CIF_ID','CUST_NAME','INTAMT_VND','LY_DO']
                df = df[cols]
                
                df_unique = df.drop_duplicates(subset='CIF_ID')
                df['INTAMT_VND'] = pd.to_numeric(df['INTAMT_VND'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                sum_intamt = df.groupby('CIF_ID', as_index=False)['INTAMT_VND'].sum()
                result = pd.merge(sum_intamt, df_unique[['CIF_ID','SOL_ID','SOL_DESC','CUST_NAME','LY_DO']], on='CIF_ID', how='left')

                # 2. Map file ngoại bảng
                dprr = read_excel_smart(file_ngoaibang)
                if "CIF 9 SO" not in dprr.columns or "SỐ TIỀN THU" not in dprr.columns:
                    st.error("❌ File ngoại bảng thiếu cột `CIF 9 SO` hoặc `Số tiền thu`!")
                    st.stop()
                dprr['SỐ TIỀN THU'] = pd.to_numeric(dprr['SỐ TIỀN THU'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                dprr_sum = dprr.groupby('CIF 9 SO', as_index=False)['SỐ TIỀN THU'].sum().rename(columns={'SỐ TIỀN THU':'Thu_ngoai_bang', 'CIF 9 SO':'CIF_ID'})
                result = result.merge(dprr_sum, on='CIF_ID', how='left')

                # 3. Map file nội bảng (Thu gốc)
                thu_goc = read_excel_smart(file_noibang)
                if "CIF 9 SO" not in thu_goc.columns or "SỐ TIỀN THU" not in thu_goc.columns:
                    st.error("❌ File nội bảng thiếu cột `CIF 9 SO` hoặc `Số tiền thu`!")
                    st.stop()
                thu_goc['SỐ TIỀN THU'] = pd.to_numeric(thu_goc['SỐ TIỀN THU'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                thu_goc_sum = thu_goc.groupby('CIF 9 SO', as_index=False)['SỐ TIỀN THU'].sum().rename(columns={'SỐ TIỀN THU':'Thu_goc_trong_thang', 'CIF 9 SO':'CIF_ID'})
                result = result.merge(thu_goc_sum, on='CIF_ID', how='left')

                # 4. Map file thu lãi
                thu_lai = read_excel_smart(file_thulai)
                if "CIF 9 SO" not in thu_lai.columns or "SỐ TIỀN THU" not in thu_lai.columns:
                    st.error("❌ File thu lãi thiếu cột `CIF 9 SO` hoặc `Số tiền thu`!")
                    st.stop()
                thu_lai['SỐ TIỀN THU'] = pd.to_numeric(thu_lai['SỐ TIỀN THU'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                thu_lai_sum = thu_lai.groupby('CIF 9 SO', as_index=False)['SỐ TIỀN THU'].sum().rename(columns={'SỐ TIỀN THU':'Thu_lai_trong_thang', 'CIF 9 SO':'CIF_ID'})
                result = result.merge(thu_lai_sum, on='CIF_ID', how='left')

                # Tính tổng thu và sinh cột theo form báo cáo định sẵn
                result['Tong_thu'] = result['Thu_goc_trong_thang'].fillna(0) + result['Thu_lai_trong_thang'].fillna(0) + result['Thu_ngoai_bang'].fillna(0)
                
                new_cols = ['Thu_thieu_lai_trong_han', 'Mien_goc', 'Mien_lai_trong_han', 'Mien_lai_qua_han', 'Mien_lai_phat', 'Ghi_chu', 'Con_ton', 'Ty_quan', 'So_thu_phi', 'note']
                for col in new_cols:
                    result[col] = 0

                result['Tham_quyen_phe_duyet'] = result['LY_DO']
                muc6 = df.groupby('CIF_ID')['INTAMT_VND'].sum().reset_index().rename(columns={'INTAMT_VND': 'muc6'})
                result = result.merge(muc6, on='CIF_ID', how='left')
                result['check_tongMGL'] = (result['Mien_goc'] + result['Mien_lai_trong_han'] + result['Mien_lai_qua_han'] + result['Mien_lai_phat']) - result['muc6']

                # Sắp xếp cấu trúc cột
                order_cols = ['SOL_ID','SOL_DESC','CIF_ID','CUST_NAME','Tong_thu','Thu_ngoai_bang','Thu_goc_trong_thang','Thu_lai_trong_thang','Thu_thieu_lai_trong_han','Mien_goc','Mien_lai_trong_han','Mien_lai_qua_han','Mien_lai_phat','Tham_quyen_phe_duyet','Ghi_chu','Con_ton','Ty_quan','So_thu_phi','muc6','check_tongMGL','note']
                result = result[order_cols]

                st.success("🎉 Tạo bảng Miễn giảm lãi thành công!")
                st.dataframe(result)
                
                excel_data = to_excel_download(result, sheet_name="XulyNo")
                st.download_button("📥 Tải Báo Cáo Xử Lý Nợ (.xlsx)", data=excel_data, file_name="XulyNo_T5_Output.xlsx")
        else:
            st.error("⚠️ Vui lòng tải lên đầy đủ các file dữ liệu phục vụ tính toán!")

# ==========================================
# TIÊU CHÍ 3: LŨY KẾ NỢ QUÁ HẠN
# ==========================================
elif menu == "Tiêu chí 3: Lũy kế nợ quá hạn":
    st.header("📌 Tiêu chí 3: Theo dõi biến động Lũy kế Nợ quá hạn qua các tháng")
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("Cấu hình chuỗi thời gian")
    thang_bat_dau = st.sidebar.text_input("Từ tháng (YYYY-MM)", "2026-01")
    thang_ket_thuc = st.sidebar.text_input("Đến tháng (YYYY-MM)", "2026-05")

    ds_thang = pd.period_range(start=thang_bat_dau, end=thang_ket_thuc, freq="M").astype(str).tolist()
    st.markdown(f"**Yêu cầu nạp file dữ liệu CRM32 cho từng tháng trong chuỗi:** `{ds_thang}`")
    
    uploaded_months_data = {}
    for m in ds_thang:
        uploaded_months_data[m] = st.file_uploader(f"Tháng {m}: Nạp file CRM32 (.xls, .xlsx)", accept_multiple_files=True, key=f"t3_{m}")

    if st.button("🚀 Chạy Tính Toán Lũy Kế"):
        all_month_data = []
        valid = True
        
        for m in ds_thang:
            if not uploaded_months_data[m]:
                st.warning(f"Chưa có tệp dữ liệu cho tháng {m}!")
                valid = False
            else:
                df_m_temp = read_excel_smart(uploaded_months_data[m], target_month=m)
                if not df_m_temp.empty:
                    all_month_data.append(df_m_temp)
                    
        if valid and len(all_month_data) > 0:
            with st.spinner("Hệ thống đang tổng hợp dữ liệu chuỗi thời gian..."):
                df = pd.concat(all_month_data, ignore_index=True)
                df.columns = df.columns.str.strip().str.upper()
                
                col_brcd, col_chi_nhanh, col_du_no, col_nhom_no, col_ngay_giai_ngan = "BRCD", "CHI_NHANH", "DU_NO_QUY_DOI", "NHOM_NO_THEO_CIF", "NGAY_GIAI_NGAN"
                
                # --- LỚP BẢO VỆ TRÁNH LỖI KEYERROR ---
                missing_cols = [c for c in [col_brcd, col_chi_nhanh, col_du_no, col_nhom_no, col_ngay_giai_ngan] if c not in df.columns]
                if missing_cols:
                    st.error(f"❌ Các file bạn upload bị thiếu hoặc sai lệch tên các cột sau: `{missing_cols}`")
                    st.markdown("💡 Danh sách tất cả các cột hệ thống hiện đọc được trong file của bạn là:")
                    st.code(list(df.columns))
                    st.stop()

                df[col_brcd] = df[col_brcd].astype(str).str.strip()
                df[col_chi_nhanh] = df[col_chi_nhanh].astype(str).str.strip()
                df[col_du_no] = pd.to_numeric(df[col_du_no], errors="coerce").fillna(0)
                df[col_nhom_no] = pd.to_numeric(df[col_nhom_no], errors="coerce")
                df[col_ngay_giai_ngan] = pd.to_datetime(df[col_ngay_giai_ngan], errors="coerce", dayfirst=True)

                ds_bao_cao_thang = []
                for m in ds_thang:
                    df_thang = df[df["THANG"] == m].copy()
                    ngay_bao_cao = pd.to_datetime(m + "-01") + pd.offsets.MonthEnd(0)
                    df_thang["SO_NGAY_TU_GIAI_NGAN_DEN_BAO_CAO"] = (ngay_bao_cao - df_thang[col_ngay_giai_ngan]).dt.days
                    df_qh = df_thang[df_thang[col_nhom_no].isin([2, 3, 4, 5])].copy()

                    giai_ngan_3_thang_qh = df_qh[(df_qh["SO_NGAY_TU_GIAI_NGAN_DEN_BAO_CAO"] >= 0) & (df_qh["SO_NGAY_TU_GIAI_NGAN_DEN_BAO_CAO"] <= 90)].groupby([col_brcd, col_chi_nhanh], as_index=False).agg(GIAI_NGAN_3_THANG_QH=(col_du_no, "sum"))
                    giai_ngan_6_thang_qh = df_qh[(df_qh["SO_NGAY_TU_GIAI_NGAN_DEN_BAO_CAO"] >= 0) & (df_qh["SO_NGAY_TU_GIAI_NGAN_DEN_BAO_CAO"] <= 180)].groupby([col_brcd, col_chi_nhanh], as_index=False).agg(GIAI_NGAN_6_THANG_QH=(col_du_no, "sum"))
                    base = df_thang.groupby([col_brcd, col_chi_nhanh], as_index=False).agg(TONG_DU_NO_THANG=(col_du_no, "sum"))

                    res_m = base.merge(giai_ngan_3_thang_qh, on=[col_brcd, col_chi_nhanh], how="left").merge(giai_ngan_6_thang_qh, on=[col_brcd, col_chi_nhanh], how="left").fillna(0)
                    res_m["THANG"] = m
                    res_m["NGAY_BAO_CAO"] = ngay_bao_cao
                    ds_bao_cao_thang.append(res_m)

                bao_cao_all_month = pd.concat(ds_bao_cao_thang, ignore_index=True)
                bao_cao_all_month["THANG_DATE"] = pd.to_datetime(bao_cao_all_month["THANG"] + "-01")
                bao_cao_all_month = bao_cao_all_month.sort_values(by=[col_brcd, "THANG_DATE"])

                # Lũy kế sửa hoàn toàn lỗi cú pháp
                bao_cao_all_month["LUY_KE_GIAI_NGAN_3_THANG_QH"] = bao_cao_all_month.groupby([col_brcd, col_chi_nhanh"])["GIAI_NGAN_3_THANG_QH"].cumsum()
                bao_cao_all_month["LUY_KE_GIAI_NGAN_6_THANG_QH"] = bao_cao_all_month.groupby([col_brcd, col_chi_nhanh"])["GIAI_NGAN_6_THANG_QH"].cumsum()

                pivot_luy_ke = bao_cao_all_month.pivot_table(index=[col_brcd, col_chi_nhanh], columns="THANG", values=["GIAI_NGAN_3_THANG_QH", "GIAI_NGAN_6_THANG_QH", "LUY_KE_GIAI_NGAN_3_THANG_QH", "LUY_KE_GIAI_NGAN_6_THANG_QH"], aggfunc="sum")
                pivot_luy_ke.columns = [f"{ct}_{th}" for ct, th in pivot_luy_ke.columns]
                pivot_luy_ke = pivot_luy_ke.reset_index()

                st.success("📊 Bảng dữ liệu biến động ngang thu được:")
                st.dataframe(pivot_luy_ke)

                out_bytes = io.BytesIO()
                with pd.ExcelWriter(out_bytes, engine="openpyxl") as wr:
                    bao_cao_all_month.to_excel(wr, sheet_name="Data luy ke theo thang", index=False)
                    pivot_luy_ke.to_excel(wr, sheet_name="Bang ngang luy ke", index=False)
                
                st.download_button("📥 Tải File Lũy Kế Đa Bản Tấm (.xlsx)", data=out_bytes.getvalue(), file_name=f"Luy_Ke_Qua_Han_{thang_bat_dau}_To_{thang_ket_thuc}.xlsx")

# =============================================================================
# TIÊU CHÍ 4: TÍN DỤNG ĐẢM BẢO BẰNG TSBD (LOGIC MỚI CỦA BẠN - CẬP NHẬT)
# =============================================================================
elif menu == "Tiêu chí 4: Tín dụng đảm bảo bằng TSBD":
    st.header("📌 Tiêu chí 4: Đối chiếu & Phân tích Dư nợ theo Loại Tài sản đảm bảo (CRM4)")
    st.info("💡 Hướng dẫn: Đăng tải trực tiếp file CRM4 của các tháng T, T-1 và tệp cấu hình mapping loại TSBD từ máy tính.")

    st.sidebar.markdown("---")
    st.sidebar.subheader("Thiết lập tháng đánh giá")
    THANG_T_1 = st.sidebar.text_input("Nhập tháng T-1 (YYYY-MM):", "2026-04")
    THANG_T = st.sidebar.text_input("Nhập tháng T (YYYY-MM):", "2026-05")

    col1, col2 = st.columns(2)
    with col1:
        files_t1 = st.file_uploader(f"Upload các file CRM4 của Tháng T-1 ({THANG_T_1}):", accept_multiple_files=True, key="t4_files_t1")
    with col2:
        files_t = st.file_uploader(f"Upload các file CRM4 của Tháng T ({THANG_T}):", accept_multiple_files=True, key="t4_files_t")
        
    st.subheader("⚙️ File cấu hình Mapping nhóm tài sản")
    mapping_file = st.file_uploader("Upload file code TSBD (.xls, .xlsx):", key="t4_mapping")

    if st.button("🚀 Chạy Đối Chiếu Tiêu Chí 4"):
        if files_t1 and files_t and mapping_file:
            with st.spinner("Hệ thống đang chạy xử lý tổng hợp dữ liệu & thiết lập báo cáo..."):
                
                # 1. Đọc file mapping nhóm tài sản
                df_map = read_excel_smart(mapping_file, auto_normalize_cols=True)
                required_map_cols = ["SECU_GROUP_CODE", "LOAI_TSBD"]
                missing_map_cols = [c for c in required_map_cols if c not in df_map.columns]
                if missing_map_cols:
                    st.error(f"❌ File mapping LOAI_TSBD thiếu cột bắt buộc: {missing_map_cols}")
                    st.stop()
                    
                df_map = df_map[["SECU_GROUP_CODE", "LOAI_TSBD"]].copy()
                df_map["SECU_GROUP_CODE"] = df_map["SECU_GROUP_CODE"].apply(clean_text)
                df_map["LOAI_TSBD"] = df_map["LOAI_TSBD"].apply(clean_text)
                df_map = df_map[df_map["SECU_GROUP_CODE"] != ""].copy()
                df_map = df_map.drop_duplicates(subset=["SECU_GROUP_CODE"], keep="first")

                # 2. Đọc file CRM4 các tháng T-1 và T
                df_t1_raw = read_excel_smart(files_t1, target_month=THANG_T_1, auto_normalize_cols=True)
                df_t_raw = read_excel_smart(files_t, target_month=THANG_T, auto_normalize_cols=True)
                
                if df_t1_raw.empty or df_t_raw.empty:
                    st.error("❌ Không thể đọc dữ liệu hoặc file upload trống!")
                    st.stop()
                    
                df_crm4 = pd.concat([df_t1_raw, df_t_raw], ignore_index=True)

                # Kiểm tra cột bắt buộc trong CRM4
                required_crm4_cols = ["BRANCH_VAY", "SECU_GROUP_CODE", "NHOM_NO", "DU_NO_QUY_DOI"]
                missing_crm4_cols = [c for c in required_crm4_cols if c not in df_crm4.columns]
                if missing_crm4_cols:
                    st.error(f"❌ Dữ liệu CRM4 thiếu các cột bắt buộc sau: {missing_crm4_cols}")
                    st.code(list(df_crm4.columns))
                    st.stop()

                # Tạo cột chi nhánh đồng bộ cấu trúc dòng lệnh
                df_crm4["BRCD"] = df_crm4["BRANCH_VAY"]
                df_crm4["CHI_NHANH"] = df_crm4["BRANCH_VAY"]

                # 3. Chuẩn hóa dữ liệu toàn bảng
                df_crm4["BRCD"] = df_crm4["BRCD"].apply(clean_text)
                df_crm4["CHI_NHANH"] = df_crm4["CHI_NHANH"].astype(str).str.strip()
                df_crm4["BRANCH_VAY"] = df_crm4["BRANCH_VAY"].apply(clean_text)
                df_crm4["SECU_GROUP_CODE"] = df_crm4["SECU_GROUP_CODE"].apply(clean_text)
                df_crm4["NHOM_NO"] = pd.to_numeric(df_crm4["NHOM_NO"], errors="coerce")
                df_crm4["DU_NO_QUY_DOI"] = df_crm4["DU_NO_QUY_DOI"].apply(clean_number).fillna(0)
                df_crm4["THANG"] = df_crm4["THANG"].astype(str).str.strip()

                df_crm4 = df_crm4[df_crm4["BRCD"] != ""].copy()

                # 4. Merge mapping nhóm tài sản bảo đảm
                df_crm4 = df_crm4.merge(df_map, on="SECU_GROUP_CODE", how="left")

                # Trích lọc lỗi để báo cáo kiểm tra chéo công tác vận hành
                df_mapping_loi = df_crm4[df_crm4["LOAI_TSBD"].isna() | (df_crm4["LOAI_TSBD"].astype(str).str.strip() == "")][
                    ["SOURCE_FILE", "THANG", "BRANCH_VAY", "BRCD", "CHI_NHANH", "SECU_GROUP_CODE", "NHOM_NO", "DU_NO_QUY_DOI"]
                ].drop_duplicates().copy()

                df_crm4["LOAI_TSBD"] = df_crm4["LOAI_TSBD"].fillna("KHAC")
                df_crm4["LOAI_TSBD"] = df_crm4["LOAI_TSBD"].apply(clean_text)

                # 5. Phân vùng chỉ lấy nhóm: HH, MMTB, PTVT
                loai_can_lay = ["HH", "MMTB", "PTVT"]
                df_tsbd = df_crm4[df_crm4["LOAI_TSBD"].isin(loai_can_lay)].copy()
                df_tsbd["IS_NO_QUA_HAN"] = df_tsbd["NHOM_NO"].isin([2, 3, 4, 5])

                # 6. Pivot Table tính tổng Dư nợ được bảo đảm
                du_no_db = df_tsbd.pivot_table(index=["THANG", "BRCD", "CHI_NHANH"], columns="LOAI_TSBD", values="DU_NO_QUY_DOI", aggfunc="sum", fill_value=0).reset_index()
                du_no_db = du_no_db.rename(columns={"HH": "DU_NO_DUOC_DAM_BAO_HH", "MMTB": "DU_NO_DUOC_DAM_BAO_MMTB", "PTVT": "DU_NO_DUOC_DAM_BAO_PTVT"})

                # 7. Pivot Table tính tổng Dư nợ quá hạn được bảo đảm
                df_qh = df_tsbd[df_tsbd["IS_NO_QUA_HAN"]].copy()
                if not df_qh.empty:
                    du_no_qh_db = df_qh.pivot_table(index=["THANG", "BRCD", "CHI_NHANH"], columns="LOAI_TSBD", values="DU_NO_QUY_DOI", aggfunc="sum", fill_value=0).reset_index()
                    du_no_qh_db = du_no_qh_db.rename(columns={"HH": "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "MMTB": "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "PTVT": "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"})
                else:
                    du_no_qh_db = pd.DataFrame(columns=["THANG", "BRCD", "CHI_NHANH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"])

                # Gộp tổng hợp chỉ tiêu theo tháng báo cáo
                tong_hop = du_no_db.merge(du_no_qh_db, on=["THANG", "BRCD", "CHI_NHANH"], how="outer").fillna(0)

                base_metric_cols = ["DU_NO_DUOC_DAM_BAO_HH", "DU_NO_DUOC_DAM_BAO_MMTB", "DU_NO_DUOC_DAM_BAO_PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"]
                for col in base_metric_cols:
                    if col not in tong_hop.columns:
                        tong_hop[col] = 0

                # Tính tỷ lệ số liệu phân vùng rủi ro
                tong_hop["TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH"] = safe_divide(tong_hop["DU_NO_QUA_HAN_DUOC_DAM_BAO_HH"], tong_hop["DU_NO_DUOC_DAM_BAO_HH"])
                tong_hop["TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB"] = safe_divide(tong_hop["DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB"], tong_hop["DU_NO_DUOC_DAM_BAO_MMTB"])
                tong_hop["TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"] = safe_divide(tong_hop["DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"], tong_hop["DU_NO_DUOC_DAM_BAO_PTVT"])

                tong_hop = tong_hop[["THANG", "BRCD", "CHI_NHANH", "DU_NO_DUOC_DAM_BAO_HH", "DU_NO_DUOC_DAM_BAO_MMTB", "DU_NO_DUOC_DAM_BAO_PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"]].copy()

                # 8. Tách chia so sánh hai thời kỳ T và T-1
                df_t1 = tong_hop[tong_hop["THANG"] == THANG_T_1].copy()
                df_t = tong_hop[tong_hop["THANG"] == THANG_T].copy()

                df_t1 = df_t1.drop(columns=["THANG"]).add_suffix("_T_1")
                df_t = df_t.drop(columns=["THANG"]).add_suffix("_T")

                bao_cao = df_t1.merge(df_t, left_on=["BRCD_T_1", "CHI_NHANH_T_1"], right_on=["BRCD_T", "CHI_NHANH_T"], how="outer")
                bao_cao["BRCD"] = bao_cao["BRCD_T"].combine_first(bao_cao["BRCD_T_1"])
                bao_cao["CHI_NHANH"] = bao_cao["CHI_NHANH_T"].combine_first(bao_cao["CHI_NHANH_T_1"])
                bao_cao = bao_cao.fillna(0)

                metrics = ["DU_NO_DUOC_DAM_BAO_HH", "DU_NO_DUOC_DAM_BAO_MMTB", "DU_NO_DUOC_DAM_BAO_PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT"]
                for m in metrics:
                    col_t1 = m + "_T_1"
                    col_t = m + "_T"
                    if col_t1 not in bao_cao.columns: bao_cao[col_t1] = 0
                    if col_t not in bao_cao.columns: bao_cao[col_t] = 0
                    bao_cao[m + "_CHENH_LECH"] = bao_cao[col_t] - bao_cao[col_t1]

                bao_cao_cols = ["BRCD", "CHI_NHANH", "DU_NO_DUOC_DAM_BAO_HH_T_1", "DU_NO_DUOC_DAM_BAO_MMTB_T_1", "DU_NO_DUOC_DAM_BAO_PTVT_T_1", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T_1", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T_1", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T_1", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T_1", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T_1", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T_1", "DU_NO_DUOC_DAM_BAO_HH_T", "DU_NO_DUOC_DAM_BAO_MMTB_T", "DU_NO_DUOC_DAM_BAO_PTVT_T", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T", "DU_NO_DUOC_DAM_BAO_HH_CHENH_LECH", "DU_NO_DUOC_DAM_BAO_MMTB_CHENH_LECH", "DU_NO_DUOC_DAM_BAO_PTVT_CHENH_LECH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_CHENH_LECH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_CHENH_LECH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_CHENH_LECH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_CHENH_LECH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_CHENH_LECH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_CHENH_LECH"]
                for col in bao_cao_cols:
                    if col not in bao_cao.columns: bao_cao[col] = 0
                bao_cao = bao_cao[bao_cao_cols].copy()

                # Đổi tên tiếng Việt
                rename_vn = {"BRCD": "BRCD", "CHI_NHANH": "CHI_NHANH", "DU_NO_DUOC_DAM_BAO_HH_T_1": "Dư nợ được đảm bảo = HH", "DU_NO_DUOC_DAM_BAO_MMTB_T_1": "Dư nợ được đảm bảo = MMTB", "DU_NO_DUOC_DAM_BAO_PTVT_T_1": "Dư nợ được đảm bảo = PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T_1": "Dư nợ quá hạn được đảm bảo = HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T_1": "Dư nợ quá hạn được đảm bảo = MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T_1": "Dư nợ quá hạn được đảm bảo = PTVT", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T_1": "Tỷ lệ Dư nợ quá hạn được đảm bảo = HH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T_1": "Tỷ lệ Dư nợ quá hạn được đảm bảo = MMTB", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T_1": "Tỷ lệ Dư nợ quá hạn được đảm bảo = PTVT", "DU_NO_DUOC_DAM_BAO_HH_T": "Dư nợ được đảm bảo = HH", "DU_NO_DUOC_DAM_BAO_MMTB_T": "Dư nợ được đảm bảo = MMTB", "DU_NO_DUOC_DAM_BAO_PTVT_T": "Dư nợ được đảm bảo = PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T": "Dư nợ quá hạn được đảm bảo = HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T": "Dư nợ quá hạn được đảm bảo = MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T": "Dư nợ quá hạn được đảm bảo = PTVT", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_T": "Tỷ lệ Dư nợ quá hạn được đảm bảo = HH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_T": "Tỷ lệ Dư nợ quá hạn được đảm bảo = MMTB", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_T": "Tỷ lệ Dư nợ quá hạn được đảm bảo = PTVT", "DU_NO_DUOC_DAM_BAO_HH_CHENH_LECH": "Dư nợ được đảm bảo = HH", "DU_NO_DUOC_DAM_BAO_MMTB_CHENH_LECH": "Dư nợ được đảm bảo = MMTB", "DU_NO_DUOC_DAM_BAO_PTVT_CHENH_LECH": "Dư nợ được đảm bảo = PTVT", "DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_CHENH_LECH": "Dư nợ quá hạn được đảm bảo = HH", "DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_CHENH_LECH": "Dư nợ quá hạn được đảm bảo = MMTB", "DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_CHENH_LECH": "Dư nợ quá hạn được đảm bảo = PTVT", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_HH_CHENH_LECH": "Tỷ lệ Dư nợ quá hạn được đảm bảo = HH", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_MMTB_CHENH_LECH": "Tỷ lệ Dư nợ quá hạn được đảm bảo = MMTB", "TY_LE_DU_NO_QUA_HAN_DUOC_DAM_BAO_PTVT_CHENH_LECH": "Tỷ lệ Dư nợ quá hạn được đảm bảo = PTVT"}
                bao_cao_vn = bao_cao.rename(columns=rename_vn).copy()

                # Sheet chi tiết đã map
                detail_cols = ["SOURCE_FILE", "THANG", "BRANCH_VAY", "BRCD", "CHI_NHANH", "CIF_KH_VAY", "TEN_KH_VAY", "LAV", "SO_KW", "CCY_KW", "SECU_SRL_NUM", "SECU_CODE", "CAP_2", "SECU_GROUP_CODE", "LOAI_TSBD", "CAP_3", "SECU_CLASS_CODE", "CAP_4", "NHOM_NO", "DU_NO", "DU_NO_QUY_DOI", "DU_NO_PHAN_BO_QUY_DOI", "SECU_VALUE", "APPORTIONED_VALUE", "APPORTIONED_VALUE_VND", "IS_NO_QUA_HAN"]
                for col in detail_cols:
                    if col not in df_tsbd.columns: df_tsbd[col] = ""
                df_detail = df_tsbd[detail_cols].copy()

                # Ghi luồng dữ liệu sang XlsxWriter để auto-styling chuẩn ngân hàng
                out_styled = io.BytesIO()
                with pd.ExcelWriter(out_styled, engine="xlsxwriter") as writer:
                    # Để dành 2 hàng đầu tiên cho việc merge nhóm lớn
                    bao_cao_vn.to_excel(writer, sheet_name="Bao_cao_T_T_1", index=False, startrow=2)
                    tong_hop.to_excel(writer, sheet_name="Tong_hop_theo_thang", index=False)
                    df_detail.to_excel(writer, sheet_name="Chi_tiet_da_map", index=False)
                    df_mapping_loi.to_excel(writer, sheet_name="Check_mapping_loi", index=False)
                    df_map.to_excel(writer, sheet_name="Mapping_LOAI_TSBD", index=False)

                    workbook  = writer.book
                    ws = writer.sheets["Bao_cao_T_T_1"]

                    header_yellow = workbook.add_format({"bold": True, "bg_color": "#FFFF00", "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True})
                    header_white = workbook.add_format({"bold": True, "bg_color": "#FFFFFF", "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True})
                    number_fmt = workbook.add_format({"num_format": "#,##0", "border": 1})
                    percent_fmt = workbook.add_format({"num_format": "0.00%", "border": 1})
                    text_fmt = workbook.add_format({"border": 1})

                    # Vẽ cấu trúc Header phức hợp nâng cao (Hàng 1, 2)
                    ws.merge_range(0, 0, 1, 0, "BRCD", header_white)
                    ws.merge_range(0, 1, 1, 1, "CHI_NHANH", header_white)
                    ws.merge_range(0, 2, 1, 2, "Dư nợ cho vay", header_yellow) # Cột C trống
                    ws.merge_range(0, 3, 0, 11, f"Tháng T-1 ({THANG_T_1})", header_yellow)
                    ws.merge_range(0, 12, 1, 12, "Dư nợ cho vay", header_yellow) # Cột M trống
                    ws.merge_range(0, 13, 0, 21, f"Tháng T ({THANG_T})", header_yellow)
                    ws.merge_range(0, 22, 0, 30, "So sánh tháng T và T-1", header_yellow)

                    # Ghi đè dòng header con số thứ tự
                    for col_num, col_name in enumerate(bao_cao_vn.columns):
                        ws.write(2, col_num, col_name, header_white)

                    ws.set_column(0, 0, 12, text_fmt)
                    ws.set_column(1, 1, 18, text_fmt)
                    ws.set_column(2, 30, 16, number_fmt)

                    # Áp dụng định dạng % cho cột tỷ lệ
                    for idx, col_name in enumerate(bao_cao_vn.columns):
                        if "Tỷ lệ" in str(col_name):
                            ws.set_column(idx, idx, 16, percent_fmt)

                    ws.autofilter(2, 0, len(bao_cao_vn) + 2, len(bao_cao_vn.columns) - 1)
                    ws.freeze_panes(3, 2)

                    # Định dạng cho các sheet phụ trợ còn lại
                    for sheet_name, df_out in [("Tong_hop_theo_thang", tong_hop), ("Chi_tiet_da_map", df_detail), ("Check_mapping_loi", df_mapping_loi), ("Mapping_LOAI_TSBD", df_map)]:
                        worksheet = writer.sheets[sheet_name]
                        for col_num, col_name in enumerate(df_out.columns):
                            worksheet.write(0, col_num, col_name, header_yellow)
                        if len(df_out.columns) > 0:
                            worksheet.autofilter(0, 0, len(df_out), len(df_out.columns) - 1)
                        worksheet.freeze_panes(1, 0)
                        
                        for i, col in enumerate(df_out.columns):
                            width = min(max(len(str(col)) + 3, 12), 40)
                            col_upper = str(col).upper()
                            if "TY_LE" in col_upper or "TỶ_LỆ" in col_upper:
                                worksheet.set_column(i, i, width, percent_fmt)
                            elif any(x in col_upper for x in ["DU_NO", "QUY_DOI", "VALUE", "APPORTIONED"]):
                                worksheet.set_column(i, i, width, number_fmt)
                            else:
                                worksheet.set_column(i, i, width, text_fmt)

                st.success(f"🎉 Hoàn thành đối chiếu chỉ tiêu! Số dòng không map được TSBD: {len(df_mapping_loi)}")
                st.write("📊 Xem trước bảng báo cáo hợp nhất:")
                st.dataframe(bao_cao_vn.head(50))
                
                st.download_button(
                    label="📥 Tải File Kết Quả Mẫu Phân Tích Mỹ Thuật (.xlsx)",
                    data=out_styled.getvalue(),
                    file_name="ket_qua_TSBD_CRM4_T4_T5.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.error("⚠️ Vui lòng cung cấp đầy đủ các file CRM4 và file Mapping nhóm tài sản!")
